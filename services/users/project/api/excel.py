from openpyxl import load_workbook
import pprint
import re

pp = pprint.PrettyPrinter(indent=4)
# from project.api.constants import DAY_TO_ISO
TEACHER = 1
TECHNICIAN = 2

DAY_TO_ISO = {"Mon": 1, "Tue": 2, "Wed": 3, "Thu": 4, "Fri": 5}


def extract_users(filename):

    try:
        wb_staff = load_workbook(filename, data_only=True)
    except IOError as e:
        return None, str(e)

    try:
        ws_teachers = wb_staff["TEACHERS"]
        ws_technicians = wb_staff["TECHNICIANS"]
    except Exception as e:
        return None, str(e)

    # check that the headers of the template are retained in uploaded file
    headers = ['Name', 'Email', 'Staff Code']

    for i in range(1, 4):
        if ws_teachers.cell(row=1, column=i).value != headers[i-1]:
            return None, 'Please ensure you use the template provided.'
    for i in range(1, 3):
        if ws_technicians.cell(row=1, column=i).value != headers[i-1]:
            return None, 'Please ensure you use the template provided.'

    staff = []

    teacher_rows = tuple(ws_teachers.rows)
    tech_rows = tuple(ws_technicians.rows)

    for i in range(2, 1 + len(teacher_rows)):
        teacher = {
            "name": ws_teachers.cell(row=i, column=1).value,
            "email": ws_teachers.cell(row=i, column=2).value,
            "role_code": TEACHER,
            "staff_code": ws_teachers.cell(row=i, column=3).value
        }
        staff.append(teacher)

    for i in range(2, 1 + len(tech_rows)):
        technician = {
            "name": ws_technicians.cell(row=i, column=1).value,
            "email": ws_technicians.cell(row=i, column=2).value,
            "role_code": TECHNICIAN,
            "staff_code": None
        }

        staff.append(technician)

    return wb_staff, staff


def extract_lessons(filename):
    wb_tt = load_workbook(filename, data_only=True)

    ws = wb_tt.active

    users_timetables = {}

    columns = tuple(ws.columns)
    rows = tuple(ws.rows)

    # get staff codes
    for i in range(2, len(columns)):
        if ws.cell(row=5, column=i).value is not None:
            users_timetables[ws.cell(row=5, column=i).value] = {}
            for j in range(6, len(rows)):
                if ws.cell(row=j, column=i).value is not None:

                    """ if lesson data is only text or full-stops, most likely
                    NOT to be a lesson... skip. """

                    if re.match('[a-z .A-Z]', ws.cell(row=j, column=i).value):
                        continue

                    users_timetables[ws.cell(
                        row=5,
                        column=i
                    ).value][ws.cell(
                        row=j,
                        column=1
                    ).value] = ws.cell(row=j, column=i).value

    lessons = []

    for staff_code in users_timetables:
        for ttlesson in users_timetables[staff_code]:

            period_full = ttlesson.split(":")

            ''' if 1 week TT, expect period format to be eg Mon:3.
            if 2 Week TT, expect period format to be eg 1Mon:3
            Test by casting first char to int. '''

            try:
                int(period_full[0][0])
                week = period_full[0][0]
                day_txt = period_full[0][1:4]
            except ValueError:
                week = 1
                day_txt = period_full[0][0:3]

            if day_txt not in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                raise ValueError('Could not parse periods.')

            # if period number is not an integer - 'reg', 'asm', 'pmr' etc,
            # skip this lesson

            try:
                period = int(period_full[1])
            except ValueError:
                continue

            class_room = users_timetables[staff_code][ttlesson].split(' ')

            lesson = {
                "staff_code": staff_code,
                "week": week,
                "period": period,
                "day_txt": day_txt,
                "day": DAY_TO_ISO[day_txt],
                "class": class_room[0],
                "room": class_room[1]
            }
            lessons.append(lesson)

    return lessons
